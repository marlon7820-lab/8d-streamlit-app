import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import datetime
import io
import os
from PIL import Image as PILImage
from io import BytesIO

# ---------------------------
# Page config
# ---------------------------
st.set_page_config(
    page_title="8D Report Assistant",
    page_icon="logo.png",
    layout="wide"
)

# ---------------------------
# App styles - updated for desktop selectbox outline + thumbnails + Root Cause textarea
# ---------------------------
st.markdown("""
<style>
/* Main app background and text */
.stApp {
    background: linear-gradient(to right, #f0f8ff, #e6f2ff);
    color: #000000 !important;
}

/* Tabs */
.stTabs [data-baseweb="tab"] {
    font-weight: bold;
    color: #000000 !important;
}

/* All textareas */
textarea {
    background-color: #ffffff !important;
    border: 1px solid #1E90FF !important;
    border-radius: 5px;
    color: #000000 !important;
}

/* Info boxes */
.stInfo {
    background-color: #e6f7ff !important;
    border-left: 5px solid #1E90FF !important;
    color: #000000 !important;
}

/* Labels */
.css-1d391kg {
    color: #1E90FF !important;
    font-weight: bold !important;
}

/* Buttons */
button[kind="primary"] {
    background-color: #87AFC7 !important;
    color: white !important;
    font-weight: bold;
}

/* Inputs, Textareas, Selectboxes styling */
div.stSelectbox, div.stTextInput, div.stTextArea {
    border: 2px solid #1E90FF !important;
    border-radius: 5px !important;
    padding: 5px !important;
    background-color: #ffffff !important;
    transition: border 0.2s ease-in-out, box-shadow 0.2s ease-in-out;
}
div.stSelectbox:hover, div.stTextInput:hover, div.stTextArea:hover {
    border: 2px solid #104E8B !important;
    box-shadow: 0 0 5px #1E90FF;
}

/* Thumbnails */
.image-thumbnail {
    width: 120px;
    height: 80px;
    object-fit: cover;
    margin: 5px;
    border: 1px solid #1E90FF;
    border-radius: 4px;
}

/* Suggesting Root Cause textarea */
.root-cause-box textarea[disabled] {
    color: #000000 !important;
    background-color: #ffffff !important;
    font-weight: bold !important;
    opacity: 1 !important;
}

/* Enable browser spellcheck and autocorrect for both English and Spanish */
textarea, input[type="text"] {
    spellcheck: true !important;
    autocorrect: on !important;
    autocapitalize: on !important;
    lang: es !important; /* Support for Spanish */
}
</style>
""", unsafe_allow_html=True)
# ---------------------------
# Reset Session check
# ---------------------------
if st.session_state.get("_reset_8d_session", False):
    preserve_keys = ["lang", "lang_key", "current_tab"]
    preserved = {k: st.session_state[k] for k in preserve_keys if k in st.session_state}
    for key in list(st.session_state.keys()):
        if key not in preserve_keys and key != "_reset_8d_session":
            del st.session_state[key]
    for k, v in preserved.items():
        st.session_state[k] = v
    st.session_state["_reset_8d_session"] = False

    # ---------------------------
    # ✅ Re-initialize 8D structure cleanly to avoid KeyErrors
    # ---------------------------
    default_template = {
        "answer": "",
        "uploaded_files": [],
        "location": [],  # empty list for multiselect
        "status": [],    # empty list for multiselect
        "occ_answer": "",
        "det_answer": "",
        "sys_answer": ""
    }

    for step in ["D1", "D2", "D3", "D4", "D5", "D6", "D7", "D8"]:
        st.session_state[step] = default_template.copy()

    # ✅ Recreate WHY lists for D5
    st.session_state["d5_occ_whys"] = []
    st.session_state["d5_det_whys"] = []
    st.session_state["d5_sys_whys"] = []
    st.rerun()

# ---------------------------
# Main title
# ---------------------------
st.markdown("<h1 style='text-align: center; color: #1E90FF;'>📋 8D Report Assistant</h1>", unsafe_allow_html=True)

# ---------------------------
# Version info
# ---------------------------
version_number = "v1.4.0"
last_updated = "October 29, 2025"
st.markdown(f"""
<hr style='border:1px solid #1E90FF; margin-top:10px; margin-bottom:5px;'>
<p style='font-size:12px; font-style:italic; text-align:center; color:#555555;'>
Version {version_number} | Last updated: {last_updated}
</p>
""", unsafe_allow_html=True)

# ---------------------------
# Sidebar: Language & Dark Mode
# ---------------------------
st.sidebar.title("8D Report Assistant")
st.sidebar.markdown("---")
st.sidebar.header("Settings")

lang = st.sidebar.selectbox("Select Language / Seleccionar Idioma", ["English", "Español"])
lang_key = "en" if lang == "English" else "es"
# ---------------------------
# Dynamic spellcheck language (English ↔ Spanish)
# ---------------------------
if lang == "English":
    spell_lang = "en"
else:
    spell_lang = "es"
dark_mode = st.sidebar.checkbox("🌙 Dark Mode")
if dark_mode:
    st.markdown("""
    <style>
    /* Main app background & text */
    .stApp {
        background: linear-gradient(to right, #1e1e1e, #2c2c2c);
        color: #f5f5f5 !important;

    /* Tabs */
    .stTabs [data-baseweb="tab"] {
        font-weight: bold; 
        color: #f5f5f5 !important;
    }
    .stTabs [data-baseweb="tab"]:hover {
        color: #87AFC7 !important;
    }

    /* Text inputs, textareas, selectboxes */
    div.stTextInput, div.stTextArea, div.stSelectbox {
        border: 2px solid #87AFC7 !important;
        border-radius: 5px !important;
        background-color: #2c2c2c !important;
        color: #f5f5f5 !important;
        padding: 5px !important;
        transition: border 0.2s ease-in-out, box-shadow 0.2s ease-in-out;
    }
    div.stTextInput:hover, div.stTextArea:hover, div.stSelectbox:hover {
        border: 2px solid #1E90FF !important;
        box-shadow: 0 0 5px #1E90FF;
    }

    /* Labels above inputs */
    div.stTextInput label,
    div.stTextArea label,
    div.stSelectbox label {
        color: #f5f5f5 !important;
        font-weight: bold;
    }

    /* Info boxes */
    .stInfo {
        background-color: #3a3a3a !important; 
        border-left: 5px solid #87AFC7 !important; 
        color: #f5f5f5 !important;
    }

    /* Sidebar background & text (kept separate) */
    .css-1d391kg {color: #87AFC7 !important; font-weight: bold !important;}
    .stSidebar {
        background-color: #1e1e1e !important;
        color: #f5f5f5 !important;
    }
    </style>
    """, unsafe_allow_html=True)

# ---------------------------
# Sidebar buttons - consistent colors in light & dark mode
# ---------------------------
st.markdown("""
<style>
/* All sidebar buttons, including Reset 8D Session & Download XLSX */
.stSidebar button,
.stSidebar .stDownloadButton button {
    background-color: #87AFC7 !important;  /* main blue */
    color: #000000 !important;             /* black text */
    font-weight: bold;
    border-radius: 5px;
    transition: background-color 0.2s ease, color 0.2s ease;
}

/* Hover effect */
.stSidebar button:hover,
.stSidebar .stDownloadButton button:hover {
    background-color: #1E90FF !important;  /* darker blue */
    color: #ffffff !important;             /* white text */
}
</style>
""", unsafe_allow_html=True)
# ---------------------------
# Step-specific guidance content (bilingual)
# ---------------------------
guidance_content = {
    "D1": {
        "en": {"title": "Define the Team & Describe the Problem","tips": """ 
- **Define the Team**:
  - Identify all team members involved in solving the issue.
  - Include functions like Quality, Engineering, Production, Supplier, etc.
  - Assign clear roles and responsibilities.
  - Example: *John (Quality) – Team Leader; Maria (Engineering) – Root Cause Analyst*.

- **Describe the Problem**:
  - Focus on **facts and measurable data** (avoid assumptions).
  - Use 5W2H (Who, What, Where, When, Why, How, How Many).
  - Example: *Customer reports radio does not power on after 2 hours of use in hot conditions*.
"""
        },
        "es": {"title": "Definir el Equipo y Describir el Problema","tips": """
- **Definir el Equipo**:
  - Identifica a todos los miembros del equipo involucrados.
  - Incluye áreas como Calidad, Ingeniería, Producción, Proveedor, etc.
  - Asigna roles y responsabilidades claras.
  - Ejemplo: *Juan (Calidad) – Líder del Equipo; María (Ingeniería) – Análisis de Causa Raíz*.

- **Describir el Problema**:
  - Enfócate en **hechos y datos medibles** (evita suposiciones).
  - Usa 5W2H (Quién, Qué, Dónde, Cuándo, Por qué, Cómo, Cuántos).
  - Ejemplo: *El cliente reporta que el radio no enciende después de 2 horas de uso en condiciones de calor*.
"""
        }
    },
    "D2": {
        "en": {"title": "Similar Parts That Could Be Affected","tips": """
- Identify parts, models, colors, or assemblies that could also be affected.
- Consider variations in suppliers, batches, or production lines.
- Example: *Front vs. rear speaker, similar model radios, alternate supplier components.*
"""
        },
        "es": {"title": "Partes Similares que Podrían Verse Afectadas","tips": """
- Identifica piezas, modelos, colores o ensamblajes que también podrían verse afectados.
- Considera variaciones de proveedores, lotes o líneas de producción.
- Ejemplo: *Altavoz delantero vs trasero, radios de modelo similar, componentes de proveedor alternativo.*
"""
        }
    },
    "D3": {
        "en": {"title": "Initial Analysis","tips": """
- Gather and review all relevant data.
- Look for patterns, trends, or unusual occurrences.
- Example: *Review production logs and defect reports to identify common failure points.*
"""
        },
        "es": {"title": "Análisis Inicial","tips": """
- Recolecta y revisa todos los datos relevantes.
- Busca patrones, tendencias o sucesos inusuales.
- Ejemplo: *Revisar registros de producción e informes de defectos para identificar puntos de falla comunes.*
"""
        }
    },
    "D4": {
        "en": {"title": "Implement Containment","tips": """
- Describe temporary actions to isolate defective material.
- Example: *Quarantined 200 pcs in warehouse, stopped shipments to customer.*
"""
        },
        "es": {"title": "Implementar Contención","tips": """
- Describe las acciones temporales para aislar material defectuoso.
- Ejemplo: *Se pusieron en cuarentena 200 piezas en almacén, se detuvieron envíos al cliente.*
"""
        }
    },
    "D5": {
        "en": {"title": "Identify Root Cause","tips": """
- Use tools like 5 Why’s or Fishbone Diagram.
- Verify the root cause with evidence.
- Example: *Incorrect torque due to missing calibration on assembly tool.*
"""
        },
        "es": {"title": "Identificar la Causa Raíz","tips": """
- Usa herramientas como 5 Porqués o Diagrama de Ishikawa.
- Verifica la causa raíz con evidencia.
- Ejemplo: *Par incorrecto debido a falta de calibración en herramienta de ensamble.*
"""
        }
    },
    "D6": {
        "en": {"title": "Verify Permanent Corrective Actions","tips": """
- Define permanent solutions to eliminate the root cause.
- Validate with testing or simulation.
- Example: *Implemented torque monitoring system to prevent missed calibrations.*
"""
        },
        "es": {"title": "Verificar Acciones Correctivas Permanentes","tips": """
- Define soluciones permanentes para eliminar la causa raíz.
- Valida con pruebas o simulaciones.
- Ejemplo: *Se implementó sistema de monitoreo de torque para evitar calibraciones omitidas.*
"""
        }
    },
    "D7": {
        "en": {"title": "Prevent Recurrence","tips": """
- Update documentation, training, and procedures.
- Example: *Updated Work Instruction #WI-321 and retrained all operators.*
"""
        },
        "es": {"title": "Prevenir Recurrencia","tips": """
- Actualiza documentación, entrenamiento y procedimientos.
- Ejemplo: *Se actualizó la Instrucción de Trabajo #WI-321 y se capacitó a todos los operadores.*
"""
        }
    },
    "D8": {
        "en": {"title": "Follow-Up Activities (Lessons Learned / Recurrence Prevention)","tips": """
- Document lessons learned from this 8D process.
- Identify opportunities to prevent similar issues in other products or lines.
- Example: *Standardized torque verification checklist applied to all new model launches.*
- Ensure sustainability of corrective actions through regular audits or reviews.
"""
        },
        "es": {"title": "Actividades de Seguimiento (Lecciones Aprendidas / Prevención de Recurrencia)","tips": """
- Documenta las lecciones aprendidas de este proceso 8D.
- Identifica oportunidades para prevenir problemas similares en otros productos o líneas.
- Ejemplo: *Lista de verificación de torque estandarizada aplicada a todos los nuevos lanzamientos de modelo.*
- Asegura la sostenibilidad de las acciones correctivas mediante auditorías o revisiones regulares.
"""
        }
    }
}

# ---------------------------
# Sidebar: App Controls
# ---------------------------
st.sidebar.markdown("---")
st.sidebar.header("⚙️ App Controls")

if st.sidebar.button("🔄 Reset 8D Session", type="primary"):
    preserve_keys = ["lang", "lang_key", "current_tab", "report_date", "prepared_by"]
    preserved = {k: st.session_state.get(k) for k in preserve_keys if k in st.session_state}
    for key in list(st.session_state.keys()):
        if key not in preserve_keys:
            del st.session_state[key]
    for k, v in preserved.items():
        st.session_state[k] = v
    st.session_state["_reset_8d_session"] = True
    st.stop()

# ---------------------------
# Language dictionary
# ---------------------------
t = {
    "en": {
        "D1": "D1: Concern Details",
        "D2": "D2: Similar Part Considerations",
        "D3": "D3: Initial Analysis",
        "D4": "D4: Implement Containment",
        "D5": "D5: Final Analysis",
        "D6": "D6: Permanent Corrective Actions",
        "D7": "D7: Countermeasure Confirmation",
        "D8": "D8: Follow-up Activities (Lessons Learned / Recurrence Prevention)",
        "Report_Date": "Report Date",
        "Prepared_By": "Prepared By",
        "Root_Cause_Occ": "Root Cause (Occurrence)",
        "Root_Cause_Det": "Root Cause (Detection)",
        "Root_Cause_Sys": "Root Cause (Systemic)",
        "Occurrence_Why": "Occurrence Why",
        "Detection_Why": "Detection Why",
        "Systemic_Why": "Systemic Why",
        "Save": "💾 Save 8D Report",
        "Download": "📥 Download XLSX",
        "Training_Guidance": "Training Guidance",
        "Example": "Example",
        "FMEA_Failure": "FMEA Failure Occurrence",
        "Location": "Material Location",
        "Status": "Activity Status",
        "Containment_Actions": "Containment Actions"
    },
    "es": {
        "D1": "D1: Detalles de la preocupación",
        "D2": "D2: Consideraciones de partes similares",
        "D3": "D3: Análisis inicial",
        "D4": "D4: Implementar contención",
        "D5": "D5: Análisis final",
        "D6": "D6: Acciones correctivas permanentes",
        "D7": "D7: Confirmación de contramedidas",
        "D8": "D8: Actividades de seguimiento (Lecciones aprendidas / Prevención de recurrencia)",
        "Report_Date": "Fecha del informe",
        "Prepared_By": "Preparado por",
        "Root_Cause_Occ": "Causa raíz (Ocurrencia)",
        "Root_Cause_Det": "Causa raíz (Detección)",
        "Root_Cause_Sys": "Causa raíz (Sistémica)",
        "Occurrence_Why": "Por qué Ocurrencia",
        "Detection_Why": "Por qué Detección",
        "Systemic_Why": "Por qué Sistémico",
        "Save": "💾 Guardar Informe 8D",
        "Download": "📥 Descargar XLSX",
        "Training_Guidance": "Guía de Entrenamiento",
        "Example": "Ejemplo",
        "FMEA_Failure": "Ocurrencia de falla FMEA",
        "Location": "Ubicación del material",
        "Status": "Estado de la actividad",
        "Containment_Actions": "Acciones de contención"
    }
}
# English
t["en"].update({
    "Concern_Details": "Concern Details",
    "Similar_Part_Considerations": "Similar Part Considerations",
    "Initial_Analysis": "Initial Analysis",
    "Follow_up_Activities": "Follow-up Activities"
})

# Spanish
t["es"].update({
    "Concern_Details": "Detalles de la Preocupación",
    "Similar_Part_Considerations": "Consideraciones de Piezas Similares",
    "Initial_Analysis": "Análisis Inicial",
    "Follow_up_Activities": "Actividades de Seguimiento"
})
# ---------------------------
# NPQP 8D steps with examples
# ---------------------------
npqp_steps = [
    ("D1", {"en":"Describe the customer concerns clearly.", "es":"Describa claramente las preocupaciones del cliente."}, {"en":"Customer reported static noise in amplifier during end-of-line test.", "es":"El cliente reportó ruido estático en el amplificador durante la prueba final."}),
    ("D2", {"en":"Check for similar parts, models, generic parts, other colors, etc.", "es":"Verifique partes similares, modelos, partes genéricas, otros colores, etc."}, {"en":"Similar model radio, Front vs. rear speaker.", "es":"Radio de modelo similar, altavoz delantero vs trasero."}),
    ("D3", {"en":"Perform an initial investigation to identify obvious issues.", "es":"Realice una investigación inicial para identificar problemas evidentes."}, {"en":"Visual inspection of solder joints, initial functional tests.", "es":"Inspección visual de soldaduras, pruebas funcionales iniciales."}),
    ("D4", {"en":"Define temporary containment actions and material location.", "es":"Defina acciones de contención temporales y ubicación del material."}, {"en":"Post Quality Alert, Increase Inspection, Inventory Certification","es":"Implementar Ayuda Visual, Incrementar Inspeccion, Certificar Inventario"}),
    ("D5", {"en": "Use 5-Why analysis to determine the root cause.", "es": "Use el análisis de 5 Porqués para determinar la causa raíz."}, {"en": "Final 'Why' from the Analysis will give a good indication of the True Root Cause", "es": "El último \"Por qué\" del análisis proporcionará una idea clara de la causa raíz del problema"}),
    ("D6", {"en":"Define corrective actions that eliminate the root cause permanently.", "es":"Defina acciones correctivas que eliminen la causa raíz permanentemente."}, {"en":"Update soldering process, redesign fixture.", "es":"Actualizar proceso de soldadura, rediseñar herramienta."}),
    ("D7", {"en":"Verify that corrective actions effectively resolve the issue.", "es":"Verifique que las acciones correctivas resuelvan efectivamente el problema."}, {"en":"Functional tests on corrected amplifiers.", "es":"Pruebas funcionales en amplificadores corregidos."}),
    ("D8", {"en":"Document lessons learned, update standards, FMEAs.", "es":"Documente lecciones aprendidas, actualice estándares, FMEAs."}, {"en":"Update SOPs, PFMEA, work instructions.", "es":"Actualizar SOPs, PFMEA, instrucciones de trabajo."})
]

# ---------------------------
# Initialize session state
# ---------------------------
for step, _, _ in npqp_steps:
    if step not in st.session_state:
        st.session_state[step] = {"answer": "", "extra": ""}
        if step in ["D1","D3","D4","D7"]:
            st.session_state[step]["uploaded_files"] = []

st.session_state.setdefault("report_date", datetime.datetime.today().strftime("%B %d, %Y"))
st.session_state.setdefault("prepared_by", "")
st.session_state.setdefault("d5_occ_whys", [""]*5)
st.session_state.setdefault("d5_det_whys", [""]*5)
st.session_state.setdefault("d5_sys_whys", [""]*5)
st.session_state.setdefault("d4_location", "")
st.session_state.setdefault("d4_status", "")
st.session_state.setdefault("d4_containment", "")

for sub in ["occ_answer", "det_answer", "sys_answer"]:
    st.session_state.setdefault(("D6"), st.session_state.get("D6", {}))
    st.session_state["D6"].setdefault(sub, "")
    st.session_state.setdefault(("D7"), st.session_state.get("D7", {}))
    st.session_state["D7"].setdefault(sub, "")
    
for step, note_dict, example_dict in npqp_steps:
    if step not in st.session_state:
        st.session_state[step] = {"answer": "", "extra": ""}
        if step in ["D1","D3","D4","D7"]:
            st.session_state[step]["uploaded_files"] = []

# ---------------------------
# Cleaned & Standardized D5 categories
# ---------------------------

# Occurrence (issues that actually happen in process, material, design, equipment, or environment)
occurrence_categories = {
    "Machine / Equipment": [
        "Equipment malfunction or inadequate maintenance",
        "Calibration drift or misalignment",
        "Tooling or fixture wear/damage",
        "Machine parameters not optimized",
        "Sensor malfunction or misalignment",
        "Process automation fault not detected",
        "Unstable process due to poor machine setup",
        "Preventive maintenance schedule not followed"
    ],
    "Material / Component": [
        "Incorrect material or component used",
        "Supplier provided off-spec component",
        "Material defect not visible during inspection",
        "Damage during storage, handling, or transport",
        "Incorrect or missing labeling / lot traceability error",
        "Material substitution without approval",
        "Material specification not aligned with requirements"
    ],
    "Process / Method": [
        "Incorrect process step sequence",
        "Inadequate process control or parameter definition",
        "Unclear or missing work instructions / procedure",
        "Process drift over time not detected",
        "Control plan not followed on production floor",
        "Incorrect torque, soldering, or assembly process",
        "Outdated or missing process FMEA linkage",
        "Process capability (Cp/Cpk) below target",
        "Lack of standardized process or method"
    ],
    "Design / Engineering": [
        "Design not robust to real-use conditions",
        "Tolerance stack-up issue not evaluated",
        "Late design change not communicated to production",
        "Incorrect or unclear drawing specification",
        "Component placement design error (DFMEA gap)",
        "Lack of design verification or validation testing"
    ],
    "Environmental / External": [
        "Temperature or humidity out of control range",
        "Electrostatic discharge (ESD) not controlled",
        "Contamination or dust affecting product",
        "Power fluctuation or interruption",
        "External vibration or noise interference",
        "Environmental monitoring process unstable"
    ]
}

# Detection (issues in QA, validation, FMEA, test setup, or organizational checks)
detection_categories = {
    "QA / Inspection": [
        "Incomplete or outdated QA checklist",
        "No automated inspection system in place",
        "Manual inspection prone to human error",
        "Inspection frequency too low to detect issue",
        "Unclear or inconsistent inspection criteria",
        "Measurement system not capable (GR&R issues)",
        "Incoming inspection missed recent supplier issue",
        "Ineffective detection method or gauge design",
        "Undefined acceptance criteria",
        "Inadequate automation or sensing",
        "Final inspection missed due to sampling plan"
    ],
    "Validation / Process": [
        "Process validation not updated after design/process change",
        "Insufficient verification of new parameters or components",
        "Design validation incomplete or not representative of real conditions",
        "Control plan coverage inadequate for potential failure modes",
        "Ongoing process monitoring missing (SPC / CpK)",
        "Containment validation ineffective",
        "Incorrect or outdated process limits"
    ],
    "FMEA / Control Plan": [
        "Failure mode not captured in PFMEA",
        "Detection controls missing or ineffective in PFMEA",
        "Control plan not updated after corrective actions",
        "FMEA not reviewed after customer complaint",
        "Detection ranking unrealistic to inspection capability",
        "PFMEA and control plan not properly linked"
    ],
    "Test / Equipment": [
        "Test equipment calibration overdue",
        "Testing software parameters incorrect",
        "Test setup cannot detect this failure mode",
        "Detection threshold too wide to capture failure",
        "Test data not logged or reviewed regularly"
    ],
    "Organizational": [
        "Feedback loop from quality incidents not implemented",
        "Weak feedback loop from Production / Quality",
        "Detection feedback missing in team meetings",
        "Incoming or in-process audit missing",
        "Training gaps in inspection/test personnel",
        "Quality alerts not properly communicated to operators"
    ]
}

# Systemic (management, training, SOPs, supplier, quality system)
systemic_categories = {
    "Management / Organization": [
        "Inadequate leadership or supervision",
        "Insufficient resource allocation",
        "Delayed response to known production issues",
        "Lack of accountability or ownership of quality issues",
        "Ineffective escalation for recurring problems",
        "Weak cross-functional communication"
    ],
    "Process / Procedure": [
        "SOPs outdated or missing",
        "Process FMEA not regularly reviewed",
        "Control plan misaligned with PFMEA or actual process",
        "Lessons learned not integrated into similar processes",
        "Inefficient document control system",
        "Preventive maintenance procedures not standardized"
    ],
    "Training": [
        "No defined training matrix or certification tracking",
        "New hires not trained on critical control points",
        "Ineffective training or onboarding process",
        "Knowledge not shared between shifts/teams",
        "Competence requirements not clearly defined"
    ],
    "Supplier / External": [
        "Supplier not included in 8D or FMEA review",
        "Supplier corrective actions not verified",
        "Incoming material audit process inadequate",
        "Supplier process changes not communicated to customer",
        "Long lead time for supplier quality issue closure",
        "Supplier violation of standards"
    ],
    "Quality System / Feedback": [
        "Internal audits ineffective or incomplete",
        "Quality KPI tracking not linked to root cause analysis",
        "Ineffective use of 5-Why or problem-solving tools",
        "Customer complaints not feeding into design reviews",
        "Lessons learned not shared or reused",
        "No systemic review after multiple 8Ds in same area"
    ]
}

occurrence_categories_es = {
    "Máquina / Equipo": [
        "Mal funcionamiento del equipo o mantenimiento inadecuado",
        "Deriva de calibración o desalineación",
        "Desgaste / daño de herramientas o accesorios",
        "Parámetros de máquina no optimizados",
        "Mal funcionamiento o desalineación del sensor",
        "Fallo en automatización del proceso no detectado",
        "Proceso inestable debido a mala configuración de la máquina",
        "Programa de mantenimiento preventivo no seguido"
    ],
    "Material / Componente": [
        "Material o componente incorrecto usado",
        "Componente fuera de especificación por proveedor",
        "Defecto de material no visible durante inspección",
        "Daño durante almacenamiento, manipulación o transporte",
        "Etiquetado incorrecto o faltante / error de trazabilidad de lote",
        "Sustitución de material sin aprobación",
        "Especificación de material no alineada con requisitos"
    ],
    "Proceso / Método": [
        "Secuencia de pasos de proceso incorrecta",
        "Control de proceso o definición de parámetros inadecuada",
        "Instrucciones de trabajo o procedimiento poco claras o faltantes",
        "Desviación del proceso no detectada con el tiempo",
        "Plan de control no seguido en producción",
        "Proceso de torque, soldadura o ensamblaje incorrecto",
        "FMEA del proceso desactualizado o faltante",
        "Capacidad del proceso (Cp/Cpk) por debajo del objetivo",
        "Falta de estandarización de proceso o método"
    ],
    "Diseño / Ingeniería": [
        "Diseño no robusto a condiciones reales",
        "Problema de acumulación de tolerancias no evaluado",
        "Cambio de diseño tardío no comunicado a producción",
        "Especificación de dibujo incorrecta o poco clara",
        "Error de colocación de componente (brecha DFMEA)",
        "Falta de verificación o validación de diseño"
    ],
    "Ambiental / Externo": [
        "Temperatura o humedad fuera del rango de control",
        "Descarga electrostática (ESD) no controlada",
        "Contaminación o polvo afectando producto",
        "Fluctuación o interrupción de energía",
        "Vibración externa o interferencia de ruido",
        "Proceso de monitoreo ambiental inestable"
    ]
}
detection_categories_es = {
    "QA / Inspección": [
        "Lista de verificación de QA incompleta o desactualizada",
        "No hay sistema de inspección automatizado",
        "Inspección manual propensa a errores humanos",
        "Frecuencia de inspección demasiado baja para detectar problemas",
        "Criterios de inspección poco claros o inconsistentes",
        "Sistema de medición no capaz (problemas GR&R)",
        "Inspección de entrada no detectó problema reciente del proveedor",
        "Método de detección o diseño de calibrador ineficaz",
        "Criterios de aceptación indefinidos",
        "Automatización o sensores inadecuados",
        "Inspección final fallida debido a plan de muestreo"
    ],
    "Validación / Proceso": [
        "Validación del proceso no actualizada tras cambio de diseño/proceso",
        "Verificación insuficiente de nuevos parámetros o componentes",
        "Validación de diseño incompleta o no representativa",
        "Cobertura del plan de control insuficiente para modos de falla potenciales",
        "Monitoreo del proceso en curso faltante (SPC / CpK)",
        "Validación de contención ineficaz",
        "Límites de proceso incorrectos o desactualizados"
    ],
    "FMEA / Plan de Control": [
        "Modo de falla no capturado en PFMEA",
        "Controles de detección faltantes o ineficaces en PFMEA",
        "Plan de control no actualizado después de acciones correctivas",
        "FMEA no revisada tras queja del cliente",
        "Clasificación de detección poco realista para la capacidad de inspección",
        "PFMEA y plan de control no correctamente vinculados"
    ],
    "Prueba / Equipos": [
        "Calibración de equipo de prueba vencida",
        "Parámetros de software de prueba incorrectos",
        "Configuración de prueba no detecta este modo de falla",
        "Umbral de detección demasiado amplio para capturar falla",
        "Datos de prueba no registrados o revisados regularmente"
    ],
    "Organizacional": [
        "Bucle de retroalimentación de incidentes de calidad no implementado",
        "Debilidad en el bucle de retroalimentación de Producción / Calidad",
        "Falta retroalimentación de detección en reuniones de equipo",
        "Auditoría de entrada o en proceso faltante",
        "Gaps de entrenamiento en personal de inspección/prueba",
        "Alertas de calidad no comunicadas correctamente a operadores"
    ]
}
systemic_categories_es = {
    "Gestión / Organización": [
        "Liderazgo o supervisión inadecuada",
        "Asignación insuficiente de recursos",
        "Respuesta retrasada a problemas de producción conocidos",
        "Falta de responsabilidad o propiedad sobre problemas de calidad",
        "Escalamiento ineficaz para problemas recurrentes",
        "Comunicación interfuncional débil"
    ],
    "Proceso / Procedimiento": [
        "SOPs desactualizados o faltantes",
        "FMEA de proceso no revisada regularmente",
        "Plan de control desalineado con PFMEA o proceso real",
        "Lecciones aprendidas no integradas en procesos similares",
        "Sistema de control de documentos ineficiente",
        "Procedimientos de mantenimiento preventivo no estandarizados"
    ],
    "Capacitación / Entrenamiento": [
        "No hay matriz de capacitación definida o seguimiento de certificaciones",
        "Nuevos empleados no entrenados en puntos críticos de control",
        "Proceso de entrenamiento o inducción ineficaz",
        "Conocimiento no compartido entre turnos/equipos",
        "Requisitos de competencia no claramente definidos"
    ],
    "Proveedor / Externo": [
        "Proveedor no incluido en revisión de 8D o FMEA",
        "Acciones correctivas de proveedor no verificadas",
        "Proceso de auditoría de material entrante inadecuado",
        "Cambios de proceso del proveedor no comunicados al cliente",
        "Tiempo de cierre de problemas de calidad del proveedor largo",
        "Proveedor violó estándares"
    ],
    "Sistema de Calidad / Retroalimentación": [
        "Auditorías internas ineficaces o incompletas",
        "Seguimiento de KPI de calidad no vinculado al análisis de causa raíz",
        "Uso ineficaz de 5-Why o herramientas de resolución de problemas",
        "Quejas de clientes no alimentan revisiones de diseño",
        "Lecciones aprendidas no compartidas o reutilizadas",
        "No hay revisión sistémica después de múltiples 8Ds en la misma área"
    ]
}

# ---------------------------
# Root cause suggestion & helper functions
# ---------------------------
def suggest_root_cause(whys, lang_key="en"):
    """
    Analyze whys (occ/det/sys) and return top 1–3 contributing root cause categories.
    Supports English and Spanish.
    """
    text = " ".join([w.lower() for w in whys if w.strip()])
    
    categories = {
        "Training / Knowledge": ["training", "knowledge", "human error", "competence", "onboarding", "guidance"],
        "Equipment / Tooling": ["equipment", "tool", "machine", "fixture", "calibration", "maintenance", "sensor"],
        "Process / Procedure": ["process", "procedure", "standard", "control plan", "method", "capability", "instructions", "fmea"],
        "Communication / Info": ["communication", "information", "handover", "feedback", "miscommunication"],
        "Material / Supplier": ["material", "supplier", "component", "part", "specification", "labeling", "lot"],
        "Design / Engineering": ["design", "specification", "drawing", "tolerance", "robust", "dfmea", "verification", "validation"],
        "Management / Resources": ["management", "supervision", "resource", "leadership", "accountability"],
        "Environment / External": ["temperature", "humidity", "contamination", "environment", "vibration", "power", "esd"]
    }

    # Count hits per category
    scores = {cat:0 for cat in categories}
    for cat, keywords in categories.items():
        for kw in keywords:
            scores[cat] += text.count(kw)
    
    scored_cats = {k:v for k,v in scores.items() if v > 0}
    if not scored_cats:
        return {
            "en": "No clear root cause suggestion (provide more detailed 5-Whys)",
            "es": "No hay sugerencia clara de causa raíz (proporcione más detalles en los 5 Porqués)"
        }[lang_key]

    # Top 3 categories
    sorted_cats = sorted(scored_cats.items(), key=lambda x: x[1], reverse=True)
    top_cats = [cat for cat, score in sorted_cats[:3]]

    # Bilingual mapping
    rc_texts = {
        "en": {
            "single": "The root cause is likely related to {0}. Focus your analysis in this area.",
            "double": "The root cause is likely related to a combination of {0} and {1}. Consider focusing your investigation in these areas.",
            "triple": "The root cause is likely related to a combination of {0}, and {1}. Focus your analysis on these areas."
        },
        "es": {
            "single": "La causa raíz probablemente está relacionada con {0}. Enfoca tu análisis en esta área.",
            "double": "La causa raíz probablemente está relacionada con una combinación de {0} y {1}. Considera enfocar tu investigación en estas áreas.",
            "triple": "La causa raíz probablemente está relacionada con una combinación de {0}, y {1}. Enfoca tu análisis en estas áreas."
        }
    }

    if len(top_cats) == 1:
        return rc_texts[lang_key]["single"].format(top_cats[0])
    elif len(top_cats) == 2:
        return rc_texts[lang_key]["double"].format(top_cats[0], top_cats[1])
    else:
        return rc_texts[lang_key]["triple"].format(top_cats[0], top_cats[1])

def render_whys_no_repeat_with_other(why_list, categories, label_prefix, lang_key="en"):
    for idx in range(len(why_list)):
        # Build options for this selectbox
        selected_so_far = [w for i, w in enumerate(why_list) if w.strip() and i != idx]
        options = [""] + [
            f"{cat}: {item}" 
            for cat, items in categories.items() 
            for item in items 
            if f"{cat}: {item}" not in selected_so_far
        ] + ["Other"]

        current_val = why_list[idx] if why_list[idx] in options else ""
        selection = st.selectbox(
            f"{label_prefix} {idx+1}",
            options,
            index=options.index(current_val) if current_val in options else 0,
            key=f"{label_prefix}_{idx}_{lang_key}"
        )

        # If "Other" is selected, show a free text box
        if selection == "Other":
            why_list[idx] = st.text_input(f"Please specify {label_prefix} {idx+1}", key=f"{label_prefix}_{idx}_other_{lang_key}")
        else:
            why_list[idx] = selection
    return why_list
# ---------------------------
# Helpers (place at top of file)
# ---------------------------
def classify_4m(text, lang="en"):
    patterns_en = {
        "Machine": ["equipment", "machine", "tool", "fixture", "wear", "maintenance", "calibration"],
        "Method": ["procedure", "process", "assembly", "sequence", "standard", "instruction", "setup"],
        "Material": ["component", "supplier", "batch", "raw", "contamination", "mix", "specification"],
        "Measurement": ["inspection", "test", "measurement", "gauge", "criteria", "frequency"]
    }
    patterns_es = {
        "Maquinaria": ["equipo", "máquina", "herramienta", "utillaje", "desgaste", "mantenimiento", "calibración"],
        "Metodo": ["procedimiento", "proceso", "ensamblaje", "secuencia", "estándar", "instrucción", "configuración"],
        "Material": ["componente", "proveedor", "lote", "materia prima", "contaminación", "mezcla", "especificación"],
        "Mediciones": ["inspección", "prueba", "medición", "calibre", "criterio", "frecuencia"]
    }
    patterns = patterns_es if lang == "es" else patterns_en
    text_lower = text.lower()
    for m, kws in patterns.items():
        if any(k in text_lower for k in kws):
            return m
    return "Other"

def smart_root_cause_suggestion(d1_concern, occ_list, det_list, sys_list, lang="en"):
    if not any([occ_list, det_list, sys_list]):
        return ("⚠️ No Why analysis provided yet.", "", "") if lang=="en" else ("⚠️ No se ha proporcionado análisis de causas.", "", "")
    
    # Suggestions dictionary
    suggestions = {
        "Method": {
            "en": ["Inadequate or missing process control or standard","Incomplete or unclear work instructions / SOPs",
                   "Outdated or obsolete process standards","Incorrect assembly or operation sequence","Missing or ineffective process controls",
                   "Lack of error-proofing (Poka-Yoke)","Variability in process execution between operators or shifts",
                   "Uncommunicated or poorly managed process changes","Process not validated or qualified"],
            "es": ["Control o estándar de proceso inadecuado o ausente","Instrucciones de trabajo / SOP incompletas o poco claras",
                   "Normas de proceso obsoletas o desactualizadas","Secuencia de montaje o operación incorrecta",
                   "Controles de proceso faltantes o ineficaces","Falta de prevención de errores (Poka-Yoke)",
                   "Variabilidad en la ejecución del proceso entre operadores o turnos","Cambios en el proceso no comunicados o mal gestionados",
                   "Proceso no validado o calificado"]
        },
        "Machine": {
            "en": ["Equipment degradation or lack of preventive maintenance","Improper machine setup or adjustment",
                   "Tooling errors (jigs, fixtures, molds)","Calibration issues","Machine design limitations",
                   "Automation or robotics malfunctions","Unstable process due to equipment variation"],
            "es": ["Degradación del equipo o falta de mantenimiento preventivo","Configuración o ajuste incorrecto de la máquina",
                   "Errores de herramientas (plantillas, fijaciones, moldes)","Problemas de calibración",
                   "Limitaciones del diseño de la máquina","Fallas en automatización o robótica",
                   "Proceso inestable debido a variación del equipo"]
        },
        "Material": {
            "en": ["Supplier or component quality variation","Incorrect material grade or specifications",
                   "Contaminated raw materials","Substandard or counterfeit components","Improper storage or handling",
                   "Material deterioration over time (aging, corrosion)","Packaging or labeling errors causing wrong part usage",
                   "Inadequate incoming inspection"],
            "es": ["Variación de calidad de proveedor o componente","Grado o especificación de material incorrecto",
                   "Materias primas contaminadas","Componentes defectuosos o falsificados","Almacenamiento o manipulación inadecuada",
                   "Deterioro del material con el tiempo (envejecimiento, corrosión)","Errores de embalaje o etiquetado causando uso incorrecto",
                   "Inspección entrante inadecuada"]
        },
        "Measurement": {
            "en": ["Insufficient inspection or gauge control","Inaccurate or uncalibrated measuring devices",
                   "Insufficient inspection frequency or sampling","Misinterpretation of measurement results",
                   "Lack of standardization in inspection procedures","Missing or incomplete measurement data",
                   "Undefined or poorly communicated tolerance limits","Measurement method not appropriate for detecting nonconformance"],
            "es": ["Inspección o control de medidores insuficiente","Dispositivos de medición inexactos o no calibrados",
                   "Frecuencia de inspección o muestreo insuficiente","Mala interpretación de los resultados de medición",
                   "Falta de estandarización en procedimientos de inspección","Datos de medición faltantes o incompletos",
                   "Límites de tolerancia mal definidos o comunicados","Método de medición no adecuado para detectar no conformidades"]
        },
        "Detection": {
            "en": ["Detection method did not identify the nonconformance before shipment",
                   "Inspection procedures not standardized or followed",
                   "Inadequate inspection frequency or sampling plan",
                   "Measurement devices not calibrated or appropriate"],
            "es": ["El método de detección no identificó la no conformidad antes del envío",
                   "Procedimientos de inspección no estandarizados o no seguidos",
                   "Frecuencia de inspección o plan de muestreo inadecuado",
                   "Dispositivos de medición no calibrados o inadecuados",
                   "Error humano durante la detección o verificación"]
        },
        "Systemic": {
            "en": ["Systemic weakness in management of change or lessons learned","Insufficient training or knowledge management",
                   "Lack of cross-functional communication","Ineffective quality management system",
                   "Inadequate corrective action follow-up or verification"],
            "es": ["Debilidad sistémica en gestión de cambios o lecciones aprendidas","Capacitación o gestión de conocimiento insuficiente",
                   "Falta de comunicación entre funciones","Sistema de gestión de calidad ineficaz",
                   "Seguimiento o verificación de acciones correctivas inadecuado"]
        },
        "Other": {
            "en": ["Perform deeper investigation","Escalate to cross-functional review"],
            "es": ["Realizar investigación más profunda","Escalar a revisión interfuncional"]
        }
    }

    occ_categories_detected = set(classify_4m(w, lang) for w in occ_list)
    occ_suggestions, det_suggestions, sys_suggestions = [], [], []

    for cat in occ_categories_detected:
        if cat in suggestions:
            occ_suggestions.extend(suggestions[cat][lang])
        else:
            occ_suggestions.extend(suggestions["Other"][lang])
    if det_list:
        det_suggestions.extend(suggestions["Detection"][lang])
    if sys_list:
        sys_suggestions.extend(suggestions["Systemic"][lang])

    # Remove duplicates
    occ_suggestions = list(dict.fromkeys(occ_suggestions))
    det_suggestions = list(dict.fromkeys(det_suggestions))
    sys_suggestions = list(dict.fromkeys(sys_suggestions))

    # Format results
    occ_result = f"💡 **Possible Occurrence Root Cause Suggestion:** {', '.join(occ_suggestions)}." if occ_suggestions else ("No Occurrence root cause detected yet." if lang=="en" else "No se detectó causa raíz de ocurrencia aún.")
    det_result = f"💡 **Possible Detection Root Cause Suggestion:** {', '.join(det_suggestions)}." if det_suggestions else ("No Detection root cause detected yet." if lang=="en" else "No se detectó causa raíz de detección aún.")
    sys_result = f"💡 **Possible Systemic Root Cause Suggestion:** {', '.join(sys_suggestions)}." if sys_suggestions else ("No Systemic root cause detected yet." if lang=="en" else "No se detectó causa raíz sistémica aún.")

    return occ_result, det_result, sys_result

# ---------------------------
# Progress tracker (NEW)
# ---------------------------
st.markdown("### 🧭 8D Completion Progress")

steps = ["D1","D2","D3","D4","D5","D6","D7","D8"]
progress = 0

d5_filled = any(w.strip() for w in st.session_state.get("d5_occ_whys", [])) \
           or any(w.strip() for w in st.session_state.get("d5_det_whys", [])) \
           or any(w.strip() for w in st.session_state.get("d5_sys_whys", []))
d6_filled = any(st.session_state.get("D6", {}).get(k, "").strip() for k in ["occ_answer","det_answer","sys_answer"])
d7_filled = any(st.session_state.get("D7", {}).get(k, "").strip() for k in ["occ_answer","det_answer","sys_answer"])
for step in steps:
    if step=="D5" and d5_filled: progress+=1
    elif step=="D6" and d6_filled: progress+=1
    elif step=="D7" and d7_filled: progress+=1
    else:
        if st.session_state.get(step, {}).get("answer", "").strip(): progress+=1

st.progress(progress/len(steps))
st.write(f"Completed {progress} of {len(steps)} steps")

# ---------------------------
# Render Tabs with Uploads
# ---------------------------
tab_labels = []
for step, _, _ in npqp_steps:
    if step == "D5":
        filled = d5_filled
    elif step == "D6":
        filled = d6_filled
    elif step == "D7":
        filled = d7_filled
    else:
        filled = st.session_state.get(step, {}).get("answer", "").strip() != ""
    
    tab_labels.append(
        f"🟢 {t[lang_key][step]}" if filled else f"🔴 {t[lang_key][step]}"
    )

tabs = st.tabs(tab_labels)

for i, (step, note_dict, example_dict) in enumerate(npqp_steps):
    with tabs[i]:
        st.markdown(f"### {t[lang_key][step]}")

        # Training Guidance & Example box
        note_text = note_dict[lang_key]
        example_text = example_dict[lang_key]
        st.markdown(f"""
<div style="
background-color:#b3e0ff;
color:black;
padding:12px;
border-left:5px solid #1E90FF;
border-radius:6px;
width:100%;
font-size:14px;
line-height:1.5;
">
<b>{t[lang_key]['Training_Guidance']}:</b> {note_text}<br><br>
💡 <b>{t[lang_key]['Example']}:</b> {example_text}
</div>
""", unsafe_allow_html=True)

        # Step-specific guidance expander from guidance_content
        gc = guidance_content[step][lang_key]
        with st.expander(f"📘 {gc['title']}"):
            st.markdown(gc["tips"])

        # File uploads for D1, D3, D4, D7
        if step in ["D1", "D3", "D4", "D7"]:
            uploaded_files = st.file_uploader(
                f"Upload files/photos for {step}",
                type=["png", "jpg", "jpeg", "pdf", "xlsx", "txt"],
                accept_multiple_files=True,
                key=f"upload_{step}"
            )
            if uploaded_files:
                for file in uploaded_files:
                    if file not in st.session_state[step]["uploaded_files"]:
                        st.session_state[step]["uploaded_files"].append(file)

            if st.session_state[step].get("uploaded_files"):
                st.markdown("**Uploaded Files / Photos:**")
                for f in st.session_state[step]["uploaded_files"]:
                    st.write(f"{f.name}")
                    if f.type.startswith("image/"):
                        st.image(f, width=192)


        # ---------------------------
        # Step-specific inputs
        # ---------------------------
        
        # D1: Customer Concern
        if step == "D1":
            st.session_state.setdefault(step, {})
            st.session_state[step]["answer"] = st.text_input(
                "Customer Concern (D1)", value=st.session_state[step].get("answer", "")
            )

        elif step == "D3":
            # D3 bilingual multiselect
            st.session_state.setdefault(step, {"inspection_stage":[]})
            options = (
                ["During Process / Manufacture", "After manufacture (e.g. Final Inspection)", "Prior dispatch"]
                if lang_key=="en"
                else ["Durante el proceso / fabricación", "Después de la fabricación (por ejemplo, inspección final)", "Antes del envío"]
            )
            st.session_state[step]["inspection_stage"] = st.multiselect(
                t[lang_key]["Inspection_Stage"] if lang_key=="en" else "Etapa de Inspección",
                options=options,
                default=st.session_state[step]["inspection_stage"]
            )

        
        elif step == "D4":
                # D4 Location / Status / Containment Actions
                st.session_state[step].setdefault("location", [])
                st.session_state[step].setdefault("status", [])
                st.session_state[step].setdefault("answer", "")

                if lang_key == "en":
                    loc_options = ["Work in progress", "Stores stock", "Warehouse stock", "Service parts"]
                    status_options = ["Pending", "In Progress", "Completed"]
                else:
                    loc_options = ["En proceso", "Stock de almacén", "Stock de bodega", "Piezas de servicio"]
                    status_options = ["Pendiente", "En Progreso", "Completado"]

                st.session_state[step]["location"] = st.multiselect(
                     t[lang_key]["Location"], options=loc_options, default=st.session_state[step]["location"]
                )
                st.session_state[step]["status"] = st.multiselect(
                    t[lang_key]["Status"], options=status_options, default=st.session_state[step]["status"]
                )
                st.session_state[step]["answer"] = st.text_area(
                    t[lang_key]["Containment_Actions"], value=st.session_state[step]["answer"], height=150
                )

            

            # ---------- D5 ----------
        elif step == "D5":
                d1_concern = st.session_state.get("D1_answer", "").strip()
                if d1_concern:
                    st.info(d1_concern)
                    st.caption("💡 Begin your Why analysis from this concern reported by the customer.")
                else:
                    st.warning("No Customer Concern defined yet in D1.")

                # Initialize whys lists in session_state if not present
                for key in ["d5_occ_whys", "d5_det_whys", "d5_sys_whys"]:
                    if key not in st.session_state:
                       st.session_state[key] = [""]


                # --- Render Occurrence / Detection / Systemic Whys ---
                if lang_key == "es":
                    st.session_state.d5_occ_whys = render_whys_no_repeat_with_other(
                        st.session_state.d5_occ_whys, occurrence_categories_es, t[lang_key]['Occurrence_Why'])
                    st.session_state.d5_det_whys = render_whys_no_repeat_with_other(
                        st.session_state.d5_det_whys, detection_categories_es, t[lang_key]['Detection_Why'])
                    st.session_state.d5_sys_whys = render_whys_no_repeat_with_other(
                        st.session_state.d5_sys_whys, systemic_categories_es, t[lang_key]['Systemic_Why'])
                else:
                    st.session_state.d5_occ_whys = render_whys_no_repeat_with_other(
                        st.session_state.d5_occ_whys, occurrence_categories, t[lang_key]['Occurrence_Why'])
                    st.session_state.d5_det_whys = render_whys_no_repeat_with_other(
                        st.session_state.d5_det_whys, detection_categories, t[lang_key]['Detection_Why'])
                    st.session_state.d5_sys_whys = render_whys_no_repeat_with_other(
                        st.session_state.d5_sys_whys, systemic_categories, t[lang_key]['Systemic_Why'])


                # --- Add buttons for extra whys ---
                if st.button("➕ Add another Occurrence Why", key=f"add_occ_{i}"):
                    st.session_state.d5_occ_whys.append("")
                if st.button("➕ Add another Detection Why", key=f"add_det_{i}"):
                    st.session_state.d5_det_whys.append("")
                if st.button("➕ Add another Systemic Why", key=f"add_sys_{i}"):
                    st.session_state.d5_sys_whys.append("")

                # --- Collect non-empty whys ---
                occ_whys = [w for w in st.session_state.d5_occ_whys if w.strip()]
                det_whys = [w for w in st.session_state.d5_det_whys if w.strip()]
                sys_whys = [w for w in st.session_state.d5_sys_whys if w.strip()]

                 # --- Duplicate check ---
                all_whys = occ_whys + det_whys + sys_whys
                duplicates = [w for w in set(all_whys) if all_whys.count(w) > 1 and w.strip()]
                if duplicates:
                    st.warning(f"⚠️ Duplicate entries detected across Occurrence/Detection/Systemic: {', '.join(duplicates)}")

                # --- Smart root cause ---
                occ_text, det_text, sys_text = smart_root_cause_suggestion(
                     d1_concern, occ_whys, det_whys, sys_whys, lang=lang_key
                )
                # --- Display results ---
                st.text_area(f"{t[lang_key]['Root_Cause_Occ']}", value=occ_text, height=120, disabled=True)
                st.text_area(f"{t[lang_key]['Root_Cause_Det']}", value=det_text, height=120, disabled=True)
                st.text_area(f"{t[lang_key]['Root_Cause_Sys']}", value=sys_text, height=120, disabled=True)

                # ---------- D6 ----------
                elif step == "D6":
                    st.session_state.setdefault("D6", {})
                    st.session_state.setdefault(step, {})
                    for sub in ["occ", "det", "sys"]:
                        key_name = f"{sub}_answer"
                        st.session_state[step].setdefault(key_name, st.session_state["D6"].get(key_name, ""))
                        st.session_state[step][key_name] = st.text_area(
                            f"D6 - Corrective Actions for {sub.capitalize()} Root Cause",
                            value=st.session_state[step][key_name],
                            key=f"d6_{sub}"
                        )
                        # ✅ store back to main D6 state so it persists
                        st.session_state["D6"][key_name] = st.session_state[step][key_name]
                            
                # ---------- D7 ----------
                elif step == "D7":
                     st.session_state.setdefault("D7", {})
                     for sub in ["occ", "det", "sys"]:
                        key_name = f"{sub}_answer"
                        st.session_state[step].setdefault(key_name, st.session_state["D7"].get(key_name, ""))
                        st.session_state[step][key_name] = st.text_area(
                             f"D7 - {sub.capitalize()} Countermeasure Verification",
                             value=st.session_state[step][key_name],
                            key=f"d7_{sub}"
                        )
                        st.session_state["D7"][key_name] = st.session_state[step][key_name]
                # ---------- D8 ----------
                elif step == "D8":
                    st.session_state.setdefault(step, {"answer": ""})
                    st.session_state[step]["answer"] = st.text_area(
                        t[lang_key]["Follow_up_Activities"],  # bilingual label
                        value=st.session_state[step]["answer"],
                        key=f"ans_{step}"
                     )

                # ---------- D2–D4 fallback ----------
                else:
                    if step not in ["D5", "D6", "D7", "D8"]:
                    # Bilingual labels for D1–D3
                        if lang_key == "es":
                            label_map = {
                                "D1": "Detalles de la Preocupación",
                                "D2": "Consideraciones de Partes Similares",
                                "D3": "Análisis Inicial",
                                "D4": "Acciones de Contención"
                             }
                        else:
                            label_map = {
                                "D1": "Concern Details",
                                "D2": "Similar Part Considerations",
                                "D3": "Initial Analysis",
                                "D4": "Containment Actions"
                            }

                        label = label_map.get(step, f"{step} – Your Answer")
                        st.session_state.setdefault(step, {"answer": ""})
                        st.session_state[step]["answer"] = st.text_area(
                            label,
                            value=st.session_state[step]["answer"],
                            key=f"ans_{step}"
                        )
   

# ---------------------------
# Collect all answers for Excel export
# ---------------------------
data_rows = []

occ_whys = [w for w in st.session_state.d5_occ_whys if w.strip()]
det_whys = [w for w in st.session_state.d5_det_whys if w.strip()]
sys_whys = [w for w in st.session_state.d5_sys_whys if w.strip()]

# --- Call the same smart bilingual root cause function used in D5 ---
if occ_whys or det_whys or sys_whys:
    occ_text, det_text, sys_text = smart_root_cause_suggestion(
        st.session_state.get("D1", {}).get("answer", ""),
        occ_whys, det_whys, sys_whys,
        lang=lang_key
    )
else:
    if lang_key == "es":
        occ_text = det_text = sys_text = "⚠️ No se ha proporcionado análisis de causas."
    else:
        occ_text = det_text = sys_text = "⚠️ No Why analysis provided yet."

# Save in session for consistency
st.session_state["D5"]["occ_root_cause"] = occ_text
st.session_state["D5"]["det_root_cause"] = det_text
st.session_state["D5"]["sys_root_cause"] = sys_text

for step, _, _ in npqp_steps:
    if step == "D1":
        # D1 text area
        answer = st.session_state[step].get("answer", "").strip()
        extra = ""  # No dropdowns in D1
        data_rows.append((step, answer, extra))
    elif step == "D2":
        # D2 text area
        answer = st.session_state[step].get("answer", "").strip()
        extra = ""  # No dropdowns
        data_rows.append((step, answer, extra))
    if step == "D6":
        data_rows.append(("D6 - Occurrence Countermeasure", st.session_state.get("D6", {}).get("occ_answer", ""), ""))
        data_rows.append(("D6 - Detection Countermeasure", st.session_state.get("D6", {}).get("det_answer", ""), ""))
        data_rows.append(("D6 - Systemic Countermeasure", st.session_state.get("D6", {}).get("sys_answer", ""), ""))
    elif step == "D7":
        data_rows.append(("D7 - Occurrence Countermeasure Verification", st.session_state.get("D7", {}).get("occ_answer", ""), ""))
        data_rows.append(("D7 - Detection Countermeasure Verification", st.session_state.get("D7", {}).get("det_answer", ""), ""))
        data_rows.append(("D7 - Systemic Countermeasure Verification", st.session_state.get("D7", {}).get("sys_answer", ""), ""))
    elif step == "D5":
        data_rows.append(("D5 - Root Cause (Occurrence)", st.session_state["D5"].get("occ_root_cause", ""), " | ".join(occ_whys)))
        data_rows.append(("D5 - Root Cause (Detection)", st.session_state["D5"].get("det_root_cause", ""), " | ".join(det_whys)))
        data_rows.append(("D5 - Root Cause (Systemic)", st.session_state["D5"].get("sys_root_cause", ""), " | ".join(sys_whys)))
    elif step == "D3":
        # ✅ Include D3 inspection stage selections in Excel export
        answer = st.session_state[step].get("answer", "")
        stages = st.session_state[step].get("inspection_stage", [])
        extra = ""
        if stages:
            label = "Inspection Stage(s)" if lang_key == "en" else "Etapa(s) de Inspección"
            extra = f"{label}: {', '.join(stages)}"
        data_rows.append((step, answer, extra))    
    elif step == "D4":
        loc_list = st.session_state[step].get("location", [])
        status_list = st.session_state[step].get("status", [])
        answer = st.session_state[step].get("answer", "")
        loc_str = ", ".join(loc_list) if loc_list else ""
        status_str = ", ".join(status_list) if status_list else ""
        extra = f"Location(s): {loc_str} | Status(es): {status_str}"
        data_rows.append((step, answer, extra))

# ---------------------------
# Excel generation function (bilingual title + color formatting)
# ---------------------------
def generate_excel():
    wb = Workbook()
    ws = wb.active

    # Bilingual worksheet title
    ws.title = "Informe 8D NPQP" if lang_key == "es" else "NPQP 8D Report"

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Add logo if exists
    if os.path.exists("logo.png"):
        try:
            img = XLImage("logo.png")
            img.width = 140
            img.height = 40
            ws.add_image(img, "A1")
        except:
            pass

    # Bilingual main title
    main_title = "📋 Asistente de Informe 8D" if lang_key == "es" else "📋 8D Report Assistant"
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    ws.cell(row=3, column=1, value=main_title).font = Font(bold=True, size=14)

    # Bilingual header info
    ws.append([t[lang_key]['Report_Date'], st.session_state.report_date])
    ws.append([t[lang_key]['Prepared_By'], st.session_state.prepared_by])
    ws.append([])

    # Header row
    header_row = ws.max_row + 1
    if lang_key == "es":
        headers = ["Etapa", "Respuesta", "Notas / Comentarios"]
    else:
        headers = ["Step", "Answer", "Extra / Notes"]

    fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=h)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Color fills for specific root cause categories
    occ_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # orange
    det_fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")  # green
    sys_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # gray

    # Bilingual keywords for color logic
    occ_keywords = ["Occurrence", "Ocurrencia"]
    det_keywords = ["Detection", "Detección"]
    sys_keywords = ["Systemic", "Sistémica"]

    # Append step answers with bilingual color formatting
    for step_label, answer_text, extra_text in data_rows:
        ws.append([step_label, answer_text, extra_text])
        r = ws.max_row
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            if c == 2:
                cell.font = Font(bold=True)
                # Apply bilingual color formatting
                if any(k in step_label for k in occ_keywords):
                    cell.fill = occ_fill
                elif any(k in step_label for k in det_keywords):
                    cell.fill = det_fill
                elif any(k in step_label for k in sys_keywords):
                    cell.fill = sys_fill

    # Insert uploaded images below table
    from PIL import Image as PILImage
    from io import BytesIO

    last_row = ws.max_row + 2
    for step in ["D1", "D3", "D4", "D7"]:
        uploaded_files = st.session_state[step].get("uploaded_files", [])
        if uploaded_files:
            title = f"{step} Archivos / Fotos Adjuntas" if lang_key == "es" else f"{step} Uploaded Files / Photos"
            ws.cell(row=last_row, column=1, value=title).font = Font(bold=True)
            last_row += 1
            for f in uploaded_files:
                if f.type.startswith("image/"):
                    try:
                        img = PILImage.open(BytesIO(f.getvalue()))
                        max_width = 300
                        ratio = max_width / img.width
                        img = img.resize((int(img.width * ratio), int(img.height * ratio)))
                        temp_path = f"/tmp/{f.name}"
                        img.save(temp_path)
                        excel_img = XLImage(temp_path)
                        ws.add_image(excel_img, f"A{last_row}")
                        last_row += int(img.height / 15) + 2
                    except Exception as e:
                        ws.cell(row=last_row, column=1, value=f"No se pudo agregar la imagen {f.name}: {e}" if lang_key == "es" else f"Could not add image {f.name}: {e}")
                        last_row += 1
                else:
                    ws.cell(row=last_row, column=1, value=f.name)
                    last_row += 1

    # Set column widths
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 60

    # ✅ Return as bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# Move download button to sidebar
with st.sidebar:
    st.download_button(
        label=t[lang_key]['Download'],
        data=generate_excel(),
        file_name=f"8D_Report_{st.session_state['report_date']}.xlsx" if lang_key == "en" else f"Informe_8D_{st.session_state['report_date']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---------------------------
# (End)
# ---------------------------
